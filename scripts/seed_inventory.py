"""
seed_inventory.py
SD_STOCK_관리표.xlsx → Supabase daily_inventory 업로드 스크립트
KP900 / CD(SD) / 2026-01-26 ~ 2026-01-30 (5일치)

실행:
  pip install openpyxl supabase python-dotenv
  python scripts/seed_inventory.py
"""

import os, re
from datetime import datetime
from openpyxl import load_workbook
from supabase import create_client
from dotenv import load_dotenv

load_dotenv()
supabase = create_client(
    os.getenv("SUPABASE_URL"),
    os.getenv("SUPABASE_SERVICE_ROLE_KEY"),
)

# ── 엑셀 파일 경로 ────────────────────────────────────────────
EXCEL_PATH = "SD_STOCK_관리표.xlsx"  # 실행 위치에 파일 놓기

# ── 시트 목록 (날짜순) ────────────────────────────────────────
DATE_SHEETS = ["1월26일", "1월27일", "1월28일", "1월29일", "1월30일"]

# ── 주령별 행 매핑 (1st half행, 2nd half행) ───────────────────
AGE_ROWS = [
    (3,   9, 10),
    (4,  12, 13),
    (5,  15, 16),
    (6,  18, 19),
    (7,  21, 22),
    (8,  24, 25),
    (9,  27, 28),
    (10, 30, None),  # 10W: half 없음
]

# ── 유틸 함수 ─────────────────────────────────────────────────
def parse_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return None

def parse_cage(v):
    """'S:10 M:20 L:71' 또는 'S이하:10' → {"S":10,"M":20,"L":71}"""
    if not v:
        return {}
    s = str(v).strip()
    result = {}
    for m in re.finditer(r"([SML])[이하:：\s]*:?\s*(\d+)", s):
        result[m.group(1)] = int(m.group(2))
    return result

def parse_record_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        m = re.search(r"(\d{4})\.\s*(\d+)\.\s*(\d+)", v)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None

def get_ids():
    """strain_id, room_id 조회"""
    strain = supabase.table("strains").select("id").eq("code", "CD(SD)").single().execute()
    room   = supabase.table("rooms").select("id").eq("room_code", "KP900").single().execute()
    return strain.data["id"], room.data["id"]

# ── 시트 파싱 ─────────────────────────────────────────────────
def parse_sheet(ws, strain_id, room_id):
    record_date = parse_record_date(ws.cell(4, 16).value)
    person      = ws.cell(4, 12).value or "JAE-MIN PARK"

    if not record_date:
        print(f"  ⚠️  날짜 파싱 실패: {ws.cell(4,16).value}")
        return []

    records = []
    for (age_week, r1, r2) in AGE_ROWS:
        halves = [("1st", r1)]
        if r2:
            halves.append(("2nd", r2))

        for (half, r) in halves:
            dob_start = parse_date(ws.cell(r, 4).value)
            dob_end   = parse_date(ws.cell(r, 6).value)

            # Male
            m_stock   = int(ws.cell(r,  7).value or 0)
            m_cage_raw= ws.cell(r,  8).value
            m_appoint = int(ws.cell(r,  9).value or 0)
            m_cut     = int(ws.cell(r, 11).value or 0)

            # Remark (행 공통 — Male/Female 공유)
            remark = str(ws.cell(r, 12).value or "").strip() or None

            # Female
            f_stock   = int(ws.cell(r, 13).value or 0)
            f_cage_raw= ws.cell(r, 14).value
            f_appoint = int(ws.cell(r, 15).value or 0)
            f_cut     = int(ws.cell(r, 17).value or 0)

            m_cage = parse_cage(m_cage_raw)
            f_cage = parse_cage(f_cage_raw)

            age_half = half if age_week < 10 else None

            # Male 레코드
            if m_stock > 0:
                records.append({
                    "record_date":        record_date,
                    "room_id":            room_id,
                    "strain_id":          strain_id,
                    "responsible_person": person,
                    "age_week":           age_week,
                    "age_half":           age_half,
                    "sex":                "M",
                    "dob_start":          dob_start,
                    "dob_end":            dob_end,
                    "total_count":        m_stock,
                    "reserved_count":     m_appoint,
                    "adjust_cut_count":   m_cut,
                    "cage_count":         sum(m_cage.values()) if m_cage else None,
                    "cage_size_breakdown":m_cage if m_cage else None,
                    "animal_type":        "retire" if age_week >= 10 else "standard",
                    "remark":             remark,
                })

            # Female 레코드
            if f_stock > 0:
                records.append({
                    "record_date":        record_date,
                    "room_id":            room_id,
                    "strain_id":          strain_id,
                    "responsible_person": person,
                    "age_week":           age_week,
                    "age_half":           age_half,
                    "sex":                "F",
                    "dob_start":          dob_start,
                    "dob_end":            dob_end,
                    "total_count":        f_stock,
                    "reserved_count":     f_appoint,
                    "adjust_cut_count":   f_cut,
                    "cage_count":         sum(f_cage.values()) if f_cage else None,
                    "cage_size_breakdown":f_cage if f_cage else None,
                    "animal_type":        "retire" if age_week >= 10 else "standard",
                    "remark":             remark,
                })

    return records

# ── 메인 ─────────────────────────────────────────────────────
def main():
    print("=== SD STOCK 재고 데이터 업로드 ===\n")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    strain_id, room_id = get_ids()
    print(f"strain_id: {strain_id}")
    print(f"room_id:   {room_id}\n")

    total_inserted = 0

    for sheet_name in DATE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  시트 없음: {sheet_name}")
            continue

        ws = wb[sheet_name]
        records = parse_sheet(ws, strain_id, room_id)

        if not records:
            print(f"  {sheet_name}: 데이터 없음")
            continue

        date = records[0]["record_date"]
        print(f"📅 {sheet_name} ({date}) — {len(records)}개 레코드 업로드 중...")

        # upsert: 같은 날짜+room+strain+age+half+sex면 덮어쓰기
        res = supabase.table("daily_inventory").upsert(
            records,
            on_conflict="record_date,room_id,strain_id,age_week,age_half,sex"
        ).execute()

        inserted = len(res.data)
        total_inserted += inserted

        # 날짜별 요약 출력
        for rec in records:
            cage_str = str(rec["cage_size_breakdown"]) if rec["cage_size_breakdown"] else "—"
            rmk = f" remark={rec['remark']}" if rec.get("remark") else ""
            print(f"  ✅ {rec['age_week']}W {rec['age_half'] or ''} {rec['sex']} | "
                  f"stock={rec['total_count']} appoint={rec['reserved_count']} "
                  f"cut={rec['adjust_cut_count']} cage={cage_str}{rmk}")

        print(f"  → {inserted}개 완료\n")

    print(f"=== 업로드 완료: 총 {total_inserted}개 레코드 ===")
    print("\n⚠️  DOB 주의:")
    print("   3W 행의 DOB가 2027년으로 파싱됨 (엑셀 수식 결과)")
    print("   실제 날짜: 2026-01-05 ~ 2026-01-07 이어야 하면 수동 확인 필요")

if __name__ == "__main__":
    main()
